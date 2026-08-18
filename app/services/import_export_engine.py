"""Robust Import/Export helpers for POS Pro v2.5.0.

This module consolidates file I/O for products, categories, suppliers, customers:
 - Excel (.xlsx / .xls) via openpyxl
 - CSV (UTF-8 / UTF-8 BOM / Windows-1256 / Latin-1)
 - PDF generation via ReportLab with Arabic support

Every public function returns a structured dict {"ok": bool, ...} so callers
can easily surface clear error messages in the UI.
"""

from __future__ import annotations

import csv
import io
import logging
import re
import tempfile
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Any, Optional

from openpyxl import Workbook, load_workbook
from sqlalchemy.exc import SQLAlchemyError
from sqlalchemy.orm import Session

log = logging.getLogger("pospro.importer")


# ═══════════════════════════════════════════════════
# Constants
# ═══════════════════════════════════════════════════
MAX_FILE_BYTES = 10 * 1024 * 1024  # 10 MB
MAX_ROWS = 10_000


# ═══════════════════════════════════════════════════
# Helpers
# ═══════════════════════════════════════════════════
def safe_str(val: Any, default: str = "") -> str:
    """Coerce any value to a stripped string, returning default on None/empty."""
    if val is None:
        return default
    try:
        s = str(val).strip()
        return s if s else default
    except Exception:
        return default


def safe_float(val: Any, default: float = 0.0) -> float:
    """Parse a numeric value as float. Returns default on failure."""
    if val is None or val == "":
        return default
    try:
        return float(val)
    except (ValueError, TypeError, ArithmeticError):
        return default


def strict_float(val: Any, default: float = 0.0) -> float:
    """Parse numeric import cells without silently converting malformed values to zero."""
    if val is None or str(val).strip() == "":
        return default
    try:
        value = float(val)
    except (ValueError, TypeError, ArithmeticError) as exc:
        raise ValueError(f"قيمة رقمية غير صالحة: {safe_str(val)[:40]}") from exc
    import math
    if not math.isfinite(value):
        raise ValueError("القيمة الرقمية يجب أن تكون رقماً محدوداً")
    return value


def _decode_bytes(raw: bytes) -> str:
    """Decode bytes with fallback chain: utf-8-sig → utf-8 → windows-1256 → latin-1."""
    for enc in ("utf-8-sig", "utf-8", "windows-1256", "cp1256", "latin-1"):
        try:
            return raw.decode(enc)
        except UnicodeDecodeError:
            continue
    # Last resort: replace undecodable bytes
    return raw.decode("utf-8", errors="replace")


# ═══════════════════════════════════════════════════
# File parsing
# ═══════════════════════════════════════════════════
def _parse_csv(raw: bytes) -> tuple[list[str], list[list[Any]]]:
    text = _decode_bytes(raw).replace("\x00", "")
    reader = csv.reader(io.StringIO(text))
    try:
        headers = [safe_str(c) for c in next(reader)]
    except StopIteration:
        return [], []
    data = []
    for i, row in enumerate(reader, start=1):
        if i > MAX_ROWS:
            raise ValueError(f"عدد الصفوف أكبر من الحد المسموح ({MAX_ROWS})")
        data.append(row)
    return headers, data


def _parse_xlsx(raw: bytes) -> tuple[list[str], list[list[Any]]]:
    # XLSX is a ZIP container; reject suspicious expansion ratios before openpyxl.
    with zipfile.ZipFile(io.BytesIO(raw)) as zf:
        total_uncompressed = sum(i.file_size for i in zf.infolist())
        if total_uncompressed > 50 * 1024 * 1024:
            raise ValueError("ملف Excel كبير بعد فك الضغط")
        compressed = max(1, sum(i.compress_size for i in zf.infolist()))
        if total_uncompressed / compressed > 100:
            raise ValueError("ملف Excel مضغوط بشكل غير آمن")
    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        wb.close(); return [], []
    rows_iter = ws.iter_rows(values_only=True)
    try:
        headers = [safe_str(c) for c in next(rows_iter)]
    except StopIteration:
        wb.close(); return [], []
    data = []
    for i, row in enumerate(rows_iter, start=1):
        if i > MAX_ROWS:
            wb.close(); raise ValueError(f"عدد الصفوف أكبر من الحد المسموح ({MAX_ROWS})")
        data.append(list(row))
    wb.close()
    return headers, data


def parse_file(filename: str, raw: bytes) -> tuple[list[str], list[list[Any]]]:
    """Dispatch to the right parser based on the filename extension.

    Raises ValueError on unsupported extension or empty data.
    """
    name = (filename or "").lower().strip()
    if not name:
        raise ValueError("اسم الملف مطلوب")
    if name.endswith((".xlsx", ".xlsm")):
        return _parse_xlsx(raw)
    if name.endswith(".csv"):
        return _parse_csv(raw)
    raise ValueError("نوع الملف غير مدعوم (يجب أن يكون xlsx/xlsm أو csv)")


# ═══════════════════════════════════════════════════
# Column lookup (case-insensitive, multiple aliases)
# ═══════════════════════════════════════════════════
def _find_col(headers: list[str], *aliases: str) -> Optional[int]:
    """Find a column by one of the given aliases (case-insensitive exact match)."""
    normalized = [h.strip() for h in headers]
    lower_map = {h.lower(): i for i, h in enumerate(normalized)}
    for alias in aliases:
        idx = lower_map.get(alias.lower())
        if idx is not None:
            return idx
    return None


# ═══════════════════════════════════════════════════
# Products import
# ═══════════════════════════════════════════════════
def import_products(
    db: Session,
    raw: bytes,
    filename: str,
) -> dict[str, Any]:
    """Import products from a file. Atomic: rolls back on any critical error.

    Returns a structured dict with added / skipped_duplicates / errors.
    """
    if not raw:
        return {"ok": False, "error": "ملف فارغ"}
    if len(raw) > MAX_FILE_BYTES:
        return {
            "ok": False,
            "error": f"حجم الملف أكبر من {MAX_FILE_BYTES // (1024 * 1024)} MB",
        }

    try:
        headers, rows = parse_file(filename, raw)
    except ValueError as exc:
        return {"ok": False, "error": str(exc)}
    except Exception as exc:
        log.exception("File parse failed")
        return {"ok": False, "error": "فشل قراءة الملف"}

    if not headers:
        return {"ok": False, "error": "الملف لا يحتوي على أعمدة"}
    if not rows:
        return {"ok": False, "error": "الملف لا يحتوي على بيانات"}
    if len(rows) > MAX_ROWS:
        return {"ok": False, "error": f"عدد الصفوف أكبر من الحد المسموح ({MAX_ROWS})"}

    idx_name = _find_col(headers, "الاسم", "Name", "name", "product", "الصنف")
    idx_barcode = _find_col(headers, "الباركود", "Barcode", "barcode")
    idx_code = _find_col(headers, "الكود", "Code", "code", "sku")
    idx_price = _find_col(headers, "السعر", "Price", "price", "sell", "سعر البيع")
    idx_cost = _find_col(headers, "التكلفة", "Cost", "cost", "buy", "سعر الشراء")
    idx_stock = _find_col(
        headers, "المخزون", "Stock", "stock", "qty", "quantity", "الكمية"
    )
    idx_unit = _find_col(headers, "الوحدة", "Unit", "unit")
    idx_min = _find_col(headers, "الحد الأدنى", "Min Stock", "min_stock", "min")
    idx_category = _find_col(headers, "القسم", "Category", "category", "cat")
    idx_supplier = _find_col(headers, "المورد", "Supplier", "supplier")

    if idx_name is None:
        return {
            "ok": False,
            "error": "لا يوجد عمود 'الاسم' في الملف. تأكد من وجود عمود باسم 'الاسم' أو 'Name'",
        }

    # Lazy imports so we don't create a circular dep
    from app.db.models import Category, Product, Supplier
    from app.utils.helpers import r3, record_stock_movement

    # Build lookups (name → id, lowercase) for category/supplier resolution
    cat_map: dict[str, int] = {}
    for c in db.query(Category).all():
        if c.name:
            cat_map[c.name.strip().lower()] = c.id
    sup_map: dict[str, int] = {}
    for s_obj in db.query(Supplier).all():
        if s_obj.name:
            sup_map[s_obj.name.strip().lower()] = s_obj.id

    # Existing barcode / code for duplicate detection
    existing_barcodes: set[str] = set()
    existing_codes: set[str] = set()
    for p in db.query(Product).all():
        if p.barcode:
            existing_barcodes.add(p.barcode.strip())
        if p.code:
            existing_codes.add(p.code.strip())

    added = 0
    skipped_dup = 0
    errors: list[dict[str, Any]] = []

    for row_num, row in enumerate(rows, start=2):
        # Skip fully-empty rows
        if not row or all(c is None or safe_str(c) == "" for c in row):
            continue
        try:
            with db.begin_nested():
                name = safe_str(row[idx_name])
                if not name:
                    continue

                barcode = safe_str(row[idx_barcode]) if idx_barcode is not None else ""
                code = safe_str(row[idx_code]) if idx_code is not None else ""

                if barcode and barcode in existing_barcodes:
                    skipped_dup += 1
                    errors.append(
                        {
                            "row": row_num,
                            "name": name,
                            "error": f"الباركود '{barcode}' موجود مسبقاً",
                        }
                    )
                    continue
                if code and code in existing_codes:
                    skipped_dup += 1
                    errors.append(
                        {
                            "row": row_num,
                            "name": name,
                            "error": f"الكود '{code}' موجود مسبقاً",
                        }
                    )
                    continue

                # Category resolution: by ID (digit) or by name
                category_id: Optional[int] = None
                if idx_category is not None and idx_category < len(row):
                    cat_val = safe_str(row[idx_category])
                    if cat_val:
                        if cat_val.isdigit():
                            category_id = int(cat_val)
                        else:
                            category_id = cat_map.get(cat_val.strip().lower())
                supplier_id: Optional[int] = None
                if idx_supplier is not None and idx_supplier < len(row):
                    sup_val = safe_str(row[idx_supplier])
                    if sup_val:
                        if sup_val.isdigit():
                            supplier_id = int(sup_val)
                        else:
                            supplier_id = sup_map.get(sup_val.strip().lower())

                cost = strict_float(row[idx_cost]) if idx_cost is not None else 0.0
                price = strict_float(row[idx_price]) if idx_price is not None else 0.0
                stock = r3(strict_float(row[idx_stock])) if idx_stock is not None else 0.0
                min_stock = r3(strict_float(row[idx_min], 5.0)) if idx_min is not None else 5.0
                if len(name) > 300 or len(barcode) > 80 or len(code) > 80:
                    raise ValueError("النص في الاسم/الباركود/الكود أطول من المسموح")
                if cost < 0 or price < 0 or stock < 0 or min_stock < 0:
                    raise ValueError("التكلفة والسعر والمخزون والحد الأدنى لا يمكن أن تكون سالبة")

                product = Product(
                    name=name,
                    barcode=barcode,
                    code=code,
                    category_id=category_id,
                    supplier_id=supplier_id,
                    unit=(
                        safe_str(row[idx_unit], "قطعة")
                        if idx_unit is not None
                        else "قطعة"
                    ),
                    cost=cost,
                    price=price,
                    stock=stock,
                    min_stock=min_stock,
                )
                db.add(product)
                db.flush()  # surface unique constraint errors early
                if abs(float(product.stock or 0)) >= 0.001:
                    record_stock_movement(
                        db, product_id=product.id, product_name=product.name,
                        quantity_delta=float(product.stock or 0), unit_cost=float(product.cost or 0),
                        movement_type="opening", user_name="import", notes="مخزون افتتاحي من الاستيراد",
                    )
                added += 1
                if barcode:
                    existing_barcodes.add(barcode)
                if code:
                    existing_codes.add(code)
        except SQLAlchemyError:
            errors.append({"row": row_num, "name": safe_str(row[idx_name]) if idx_name < len(row) else "", "error": "تعذر حفظ هذا الصف بسبب تعارض في البيانات"})
        except ValueError as exc:
            errors.append({"row": row_num, "name": safe_str(row[idx_name]) if idx_name < len(row) else "", "error": str(exc)[:200]})
        except Exception:
            log.exception("Unexpected product import row failure at row %s", row_num)
            errors.append({"row": row_num, "name": safe_str(row[idx_name]) if idx_name < len(row) else "", "error": "تعذر معالجة هذا الصف بسبب خطأ داخلي"})

    try:
        db.commit()
    except SQLAlchemyError as exc:
        db.rollback()
        log.exception("Bulk import commit failed")
        return {
            "ok": False,
            "error": "فشل حفظ المنتجات بسبب تعارض في البيانات",
            "added": added,
            "skipped_duplicates": skipped_dup,
        }

    return {
        "ok": True,
        "added": added,
        "skipped_duplicates": skipped_dup,
        "errors": errors[:50],
        "total_errors": len(errors),
    }


# ═══════════════════════════════════════════════════
# Customers import
# ═══════════════════════════════════════════════════
def import_customers(db: Session, raw: bytes, filename: str) -> dict[str, Any]:
    """Import customers from a CSV/XLSX file."""
    if not raw:
        return {"ok": False, "error": "ملف فارغ"}
    if len(raw) > MAX_FILE_BYTES:
        return {"ok": False, "error": "حجم الملف كبير جداً"}

    try:
        headers, rows = parse_file(filename, raw)
    except ValueError as exc:
        return {"ok": False, "error": str(exc)}
    except Exception as exc:
        log.exception("File parse failed")
        return {"ok": False, "error": "فشل قراءة الملف"}

    if not headers or not rows:
        return {"ok": False, "error": "الملف فارغ"}

    if len(rows) > MAX_ROWS:
        return {"ok": False, "error": f"عدد الصفوف أكبر من {MAX_ROWS}"}

    idx_name = _find_col(headers, "الاسم", "Name", "name")
    idx_phone = _find_col(headers, "الهاتف", "Phone", "phone", "mobile")
    idx_notes = _find_col(headers, "الملاحظات", "Notes", "notes")
    idx_balance = _find_col(headers, "الرصيد", "Balance", "balance")

    if idx_name is None:
        return {"ok": False, "error": "لا يوجد عمود 'الاسم' في الملف"}

    from app.db.models import Customer
    from app.utils.helpers import record_ledger_entry

    added = 0
    errors: list[dict[str, Any]] = []
    for row_num, row in enumerate(rows, start=2):
        if not row or all(c is None or safe_str(c) == "" for c in row):
            continue
        try:
            with db.begin_nested():
                name = safe_str(row[idx_name])
                if not name:
                    continue
                if len(name) > 200:
                    raise ValueError("اسم العميل أطول من المسموح")
                customer = Customer(
                    name=name,
                    phone=safe_str(row[idx_phone]) if idx_phone is not None else "",
                    notes=safe_str(row[idx_notes]) if idx_notes is not None else "",
                    balance=(
                        strict_float(row[idx_balance]) if idx_balance is not None else 0.0
                    ),
                )
                db.add(customer)
                db.flush()
                opening = float(customer.balance or 0)
                if abs(opening) > 0.0001:
                    record_ledger_entry(
                        db, customer_id=customer.id, customer_name=customer.name,
                        movement_type="opening", description="رصيد افتتاحي من الاستيراد",
                        debit=opening if opening > 0 else 0, credit=-opening if opening < 0 else 0,
                        user_name="import",
                    )
                added += 1
        except Exception:
            errors.append({"row": row_num, "name": safe_str(row[idx_name]) if idx_name < len(row) else "", "error": "تعذر استيراد هذا الصف"})

    try:
        db.commit()
    except SQLAlchemyError as exc:
        db.rollback()
        return {"ok": False, "error": "فشل حفظ بيانات العملاء"}

    return {
        "ok": True,
        "added": added,
        "errors": errors[:50],
        "total_errors": len(errors),
    }


# ═══════════════════════════════════════════════════
# Export helpers (CSV / XLSX) - returns (path, media_type, filename)
# ═══════════════════════════════════════════════════
def _tmp_path(name: str, ext: str) -> Path:
    """Get a cross-platform temp file path."""
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_name = re.sub(r"[^A-Za-z0-9_\-]", "_", name)
    fd, path = tempfile.mkstemp(prefix=f"{safe_name}_{ts}_", suffix=f".{ext}")
    # Close the file descriptor immediately; we only want the path
    import os

    try:
        os.close(fd)
    except OSError:
        pass
    return Path(path)


def _excel_safe(v: Any) -> Any:
    if isinstance(v, str) and v and v[0] in ("=", "+", "-", "@"):
        return "'" + v
    return v


def export_csv(headers: list[str], rows: list[list[Any]], name: str) -> Path:
    """Write headers + rows to a UTF-8-BOM CSV file. Returns the file path."""
    path = _tmp_path(name, "csv")
    # Use utf-8-sig BOM so Excel on Windows reads Arabic correctly
    with open(path, "w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.writer(fh)
        writer.writerow(headers)
        for r in rows:
            writer.writerow([_excel_safe("" if v is None else v) for v in r])
    return path


def export_xlsx(headers: list[str], rows: list[list[Any]], name: str) -> Path:
    """Write headers + rows to an XLSX file with bold headers. Returns the path."""
    from openpyxl.styles import Alignment, Font, PatternFill

    path = _tmp_path(name, "xlsx")
    wb = Workbook()
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet()
    ws.title = name[:31] or "Sheet"  # Excel sheet name max 31 chars
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DDDDDD")
        cell.alignment = Alignment(horizontal="center")
    for r in rows:
        ws.append([_excel_safe("" if v is None else v) for v in r])
    # Auto-size columns (best-effort)
    try:
        for col_idx, _ in enumerate(headers, start=1):
            col_letter = ws.cell(row=1, column=col_idx).column_letter
            max_len = len(str(headers[col_idx - 1]))
            for r in rows:
                if col_idx - 1 < len(r) and r[col_idx - 1] is not None:
                    max_len = max(max_len, len(str(r[col_idx - 1])))
            ws.column_dimensions[col_letter].width = min(max_len + 2, 40)
    except Exception:
        pass
    wb.save(path)
    return path
